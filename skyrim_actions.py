# skyrim_actions.py
from __future__ import annotations
import os
from itertools import combinations
from collections import defaultdict
import pandas as pd
from ortools.sat.python import cp_model
import time
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


def run_min_cover_actions(csv_path: str, out_dir: str | None = None) -> str:
    """
    Runs the potions+eat minimum cover optimization.
    Returns the path to the output CSV written.
    """

    class ProgressPrinter(cp_model.CpSolverSolutionCallback):
        def __init__(self, every_seconds: float = 2.0, stop_if_no_improve_s: float = 120.0):
            super().__init__()
            self.every_seconds = every_seconds
            self.stop_if_no_improve_s = stop_if_no_improve_s

            self._last_print = time.time()
            self._best = None
            self._last_improve_time = time.time()

        def OnSolutionCallback(self):
            now = time.time()
            obj = self.ObjectiveValue()

            # detect improvement
            if (self._best is None) or (obj < self._best):
                self._best = obj
                self._last_improve_time = now
                print(f"[CP-SAT] IMPROVED: best={obj:.0f} at t={self.WallTime():.1f}s")

            # periodic heartbeat print
            if now - self._last_print >= self.every_seconds:
                self._last_print = now
                print(
                    f"[CP-SAT] t={self.WallTime():.1f}s "
                    f"best={self._best:.0f} "
                    f"since_improve={now - self._last_improve_time:.0f}s "
                    f"conflicts={self.NumConflicts()} branches={self.NumBranches()}"
                )

            # stop if no improvement for too long
            if now - self._last_improve_time >= self.stop_if_no_improve_s:
                print(f"[CP-SAT] Stopping: no improvement in {self.stop_if_no_improve_s:.0f}s.")
                self.StopSearch()


    if out_dir is None:
        out_dir = os.path.dirname(os.path.abspath(csv_path)) or "."

    df = pd.read_csv(csv_path)
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Build data structures
    effects = {}   # ingredient -> set(all effects incl effect0)
    primary = {}   # ingredient -> effect0
    for _, row in df.iterrows():
        ing = row["ingredient"]

        e0 = row["effect0"]
        e1 = row["effect1"]
        e2 = row["effect2"]
        e3 = row["effect3"]

        # normalize / drop blanks
        all_effs = [e0, e1, e2, e3]
        all_effs = [e for e in all_effs if isinstance(e, str) and e != ""]

        effects[ing] = set(all_effs)

        if isinstance(e0, str) and e0 != "":
            primary[ing] = e0
        else:
            raise ValueError(f"{ing} is missing effect0 (primary).")

    ingredients = list(effects.keys())
    print("N ingredients:", len(ingredients))

    # effect -> ingredients index (for potion generation)
    effect_to_ings = defaultdict(list)
    for ing, effs in effects.items():
        for e in effs:
            effect_to_ings[e].append(ing)

    # Candidate potions: pairs
    pair_to_shared = {}  # (i,j) -> set(shared effects)
    for e, ings in effect_to_ings.items():
        for a, b in combinations(sorted(ings), 2):
            key = (a, b)
            pair_to_shared.setdefault(key, set()).add(e)

    valid_pairs = [(a, b, shared) for (a, b), shared in pair_to_shared.items()]
    print("Valid pairs:", len(valid_pairs))

    def potion_effects(triple):
        """Effects that would actually appear in the potion (present in >=2 ingredients)."""
        a, b, c = triple
        ea, eb, ec = effects[a], effects[b], effects[c]
        return (ea & eb) | (ea & ec) | (eb & ec)

    # neighbors for triple generation
    neighbors = defaultdict(set)
    for a, b, shared in valid_pairs:
        neighbors[a].add(b)
        neighbors[b].add(a)

    # Candidate potions: useful triples
    useful_triples = []
    seen = set()
    for a in ingredients:
        nbrs = sorted(neighbors[a])
        for b, c in combinations(nbrs, 2):
            triple = tuple(sorted((a, b, c)))
            if triple in seen:
                continue
            seen.add(triple)
            pe = potion_effects(triple)
            if len(pe) >= 2:  # keep your usefulness threshold
                useful_triples.append((triple[0], triple[1], triple[2], pe))

    print("Useful triples:", len(useful_triples))

    # Universe U: all (ingredient, effect) pairs to learn
    U = [(ing, eff) for ing, effs in effects.items() for eff in effs]
    u_index = {pair: k for k, pair in enumerate(U)}
    m = len(U)
    print("Universe size |U| =", m)

    def coverage_for_potion(ings, potion_effects_set):
        """Universe indices covered by crafting potion with these ingredients."""
        covered = set()
        for ing in ings:
            for e in potion_effects_set:
                if e in effects[ing]:
                    covered.add(u_index[(ing, e)])
        return covered

    # Build candidates list
    candidates = []

    # Potion candidates from pairs
    for a, b, shared in valid_pairs:
        cov = coverage_for_potion((a, b), shared)
        if cov:
            candidates.append(("potion", (a, b), shared, cov))

    # Potion candidates from triples
    for a, b, c, pe in useful_triples:
        cov = coverage_for_potion((a, b, c), pe)
        if cov:
            candidates.append(("potion", (a, b, c), pe, cov))

    # Eat candidates (covers only (ing, effect0))
    for ing, e0 in primary.items():
        cov = {u_index[(ing, e0)]}
        candidates.append(("eat", (ing,), None, cov))

    print("Total candidates (potions + eat):", len(candidates))

    # Greedy baseline
    def greedy_set_cover(candidates, m):
        uncovered = set(range(m))
        chosen = []
        while uncovered:
            best_idx = None
            best_gain = 0
            for idx, (_, _, _, cov) in enumerate(candidates):
                gain = len(cov & uncovered)
                if gain > best_gain:
                    best_gain = gain
                    best_idx = idx
            if best_gain == 0:
                raise RuntimeError("Greedy got stuck: some universe items are not coverable by your candidates.")
            chosen.append(best_idx)
            uncovered -= candidates[best_idx][3]
        return chosen

    greedy_choice = greedy_set_cover(candidates, m)
    print("Greedy uses", len(greedy_choice), "actions (potions + eat)")

    # Incidence lists
    covers_u = [[] for _ in range(m)]
    for j, (_, _, _, cov) in enumerate(candidates):
        for u in cov:
            covers_u[u].append(j)

    uncoverable = [u for u in range(m) if not covers_u[u]]
    if uncoverable:
        bad = [U[u] for u in uncoverable[:20]]
        raise RuntimeError(f"Uncoverable universe items (showing up to 20): {bad}")

    # CP-SAT solve
    model = cp_model.CpModel()
    x = [model.NewBoolVar(f"x_{j}") for j in range(len(candidates))]

    for u in range(m):
        model.Add(sum(x[j] for j in covers_u[u]) >= 1)

    # Minimize total actions (each potion or eat counts as 1)
    model.Minimize(sum(x))

    # Hint with greedy
    greedy_set = set(greedy_choice)
    for j in range(len(candidates)):
        model.AddHint(x[j], 1 if j in greedy_set else 0)

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = 8
    solver.parameters.max_time_in_seconds = 240.0

    print("Starting CP-SAT search...")
    cb = ProgressPrinter(every_seconds=2.0, stop_if_no_improve_s=120.0)

    # Version-tolerant solve call:
    if hasattr(solver, "SolveWithSolutionCallback"):
        result = solver.SolveWithSolutionCallback(model, cb)
    else:
        # Newer OR-Tools: callback is passed to Solve/solve (argument name varies by version)
        try:
            result = solver.Solve(model, cb)  # sometimes positional works
        except TypeError:
            try:
                result = solver.Solve(model, solution_callback=cb)
            except TypeError:
                # Some versions prefer lower-case solve() and may use 'callback' instead
                try:
                    result = solver.solve(model, solution_callback=cb)
                except TypeError:
                    result = solver.solve(model, callback=cb)

    print("Status:", solver.StatusName(result))
    if result not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("No solution found (unexpected if candidates cover everything).")

    chosen = [j for j in range(len(candidates)) if solver.Value(x[j]) == 1]
    print("Exact solution uses", len(chosen), "actions (potions + eat)")

    # Output
    rows = []

    for j in chosen:
        kind, ings, pe_set, cov = candidates[j]

        # group learned effects by ingredient
        by_ing = defaultdict(list)
        for u in cov:
            ing, eff = U[u]
            by_ing[ing].append(eff)

        if kind == "eat":
            action_str = f"EAT {ings[0]}"
            potion_eff_str = ""
        else:
            action_str = " + ".join(ings)
            potion_eff_str = ", ".join(sorted(pe_set))

        rows.append({
            "type": kind,  # "potion" or "eat"
            "action": action_str,
            "num_pairs_learned": len(cov),
            "potion_effects": potion_eff_str,
            "effects_by_ingredient": "; ".join(
                f"{ing}: {', '.join(sorted(set(by_ing[ing])))}" for ing in ings
            ),
        })

    # Build the DataFrame once
    out_df = pd.DataFrame(rows).sort_values(
        ["type", "num_pairs_learned"], ascending=[True, False]
    )

    # Save the CSV (plain)
    # csv_path = os.path.join(out_dir, "minimum_cover_actions_readable.csv")
    # out_df.to_csv(csv_path, index=False)
    # print("Wrote", csv_path)

    # Save an Excel workbook (formatted + 2 sheets)
    xlsx_path = os.path.join(out_dir, "minimum_cover_actions.xlsx")
    write_actions_excel(out_df.to_dict("records"), xlsx_path)
    print("Wrote", xlsx_path)

    return xlsx_path



def write_actions_excel(rows: list[dict], out_xlsx_path: str) -> None:
    """
    rows: list of dicts with keys:
      type, action, num_pairs_learned, potion_effects, effects_by_ingredient
    Creates an .xlsx with:
      Sheet 'Craft List' with checkbox + green row highlight, hidden columns, auto-fit widths.
      Sheet 'Ingredient Totals' with ingredient counts used in craft actions.
    """
    wb = Workbook()

    # -------------------------
    # Sheet 1: Craft List
    # -------------------------
    ws = wb.active
    ws.title = "Craft List"

    # We'll create a user-friendly sheet:
    # A: Done (checkbox)
    # B: Craft (action)
    # (hide the other columns by default)
    headers = ["Done", "Craft", "type", "num_pairs_learned", "potion_effects", "effects_by_ingredient"]
    ws.append(headers)

    # Write rows
    for r in rows:
        ws.append([
            False,                 # Done checkbox default unchecked
            r["action"],           # Craft
            r["type"],
            r["num_pairs_learned"],
            r["potion_effects"],
            r["effects_by_ingredient"],
        ])

    n_rows = ws.max_row

    # Hide columns that aren't relevant for end users:
    # C, D, E, F correspond to type, num_pairs_learned, potion_effects, effects_by_ingredient
    for col_letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].hidden = True

    # Add checkbox validation to column A (Done)
    # In Excel, a TRUE/FALSE dropdown is the most compatible "checkbox-like" approach.
    # If you truly need clickable checkbox objects, that's much harder and less portable.
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"A2:A{n_rows}")

    # Conditional formatting: if Done is TRUE, fill the row green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # Apply to columns A..B (visible columns) across the whole used range.
    # If you want it to color hidden columns too, change "A".."B" to "A".."F".
    rule = FormulaRule(formula=['$A2=TRUE'], fill=green_fill)
    ws.conditional_formatting.add(f"A2:B{n_rows}", rule)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-fit column widths (for visible columns; you can include all if you want)
    for col_idx in range(1, 3):  # A..B
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # -------------------------
    # Sheet 2: Ingredient Totals
    # -------------------------
    ws2 = wb.create_sheet("Ingredient Totals")
    ws2.append(["Ingredient", "Count"])

    # Count ingredients used in potion actions (and optionally eat actions)
    counts = Counter()

    for r in rows:
        action = r["action"]
        typ = r["type"]
        if typ == "potion":
            # action looks like "A + B" or "A + B + C"
            parts = [p.strip() for p in action.split("+")]
            for p in parts:
                if p:
                    counts[p] += 1
        elif typ == "eat":
            # If you want eats included in totals, uncomment this:
            # action looks like "EAT Ingredient"
            # ing = action.replace("EAT", "", 1).strip()
            # if ing:
            #     counts[ing] += 1
            pass

    for ing, ct in counts.most_common():
        ws2.append([ing, ct])

    ws2.freeze_panes = "A2"

    # Auto-fit Ingredient Totals columns
    for col_idx in range(1, 3):  # A..B
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws2[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(out_xlsx_path)